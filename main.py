from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__)

products: list = []
suppliers: list = []
products_with_supplier: dict = {}

def init():
    wb = openpyxl.load_workbook('data\\Model_zakupa_dlya_khakatona.xlsx', read_only=True)
    products_sheet = wb["список товаров"]
    suppliers_sheet = wb["Наименования поставщика"]
    ws = wb['Область работы']

    global products
    #data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    products = []
    for row in products_sheet.iter_rows(min_row=2, values_only=True):
        category_3, category_nom, element_nom, total, *_ = row
        if element_nom == None:
            continue
        products.append((category_3, category_nom, element_nom, total))

    global suppliers
    suppliers = []
    for row in suppliers_sheet.iter_rows(min_row=2, values_only=True):
        lol, name, availability, supplier, *_ = row
        suppliers.append((name, availability, supplier))

    global products_with_supplier
    #print(suppliers)
    products_with_supplier = {}
    for product in products:
        product_name = product[2]
        for supplier in suppliers:
            if product_name in supplier:
                if product_name in products_with_supplier:
                    products_with_supplier[product_name].append(supplier[2])
                else:
                    products_with_supplier[product_name] = [supplier[2]]

#default login/reister page
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add', methods=['GET', 'POST'])
def add():
    if request.method == 'POST':
        pass
    elif request.method == 'GET':
        return render_template('add.html')

@app.route('/aa')
def aa():
    return render_template('aaadd.html')

@app.route('/product')
def product():
    global products
    return render_template('xslx_test.html', data=products)

@app.route('/prod_sup')
def parsed():
    global products_with_supplier
    return render_template('pr_w_sup.html', products=products_with_supplier)

init()
if __name__ == '__main__':
    app.run(debug=True)
